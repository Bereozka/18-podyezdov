from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from typing import List
from typing_extensions import Literal


def create_excel_file(
    work_data: list,
    material_data: list,
    template_path: str,
    final_file_path: str,
) -> None:
    """Adds data from the transmitted data to the template

    Parameters
    ----------
    work_data : list
        Data to be entered in the first worksheet
    material_data : list
        Data to be entered in the second worksheet
    template_path : str
        Path to the template
    final_file_path : str
        The path to the final file

    Raises
    ------
    ValueError
        If one of the parameters does not match its type

    """

    if not isinstance(work_data, list):
        raise ValueError("Parameter work_data should be a list")
    elif not isinstance(material_data, list):
        raise ValueError("Parameter material_data should be a list")
    elif not isinstance(template_path, str):
        raise ValueError("Parameter template_path should be a str")
    elif not isinstance(final_file_path, str):
        raise ValueError("Parameter final_file_path should be a str")

    workbook = load_workbook(template_path)

    sheets = workbook.sheetnames

    material_worksheet = workbook[sheets[1]]

    worksheet = workbook.active
    insert_rows(
        worksheet,
        work_data,
        list(range(2, len(work_data) + 2)),
    )

    insert_rows(
        material_worksheet,
        material_data,
        list(range(2, len(material_data) + 2)),
    )

    workbook.save(final_file_path)


def insert_rows(
    worksheet: Worksheet,
    data: List[dict],
    numbers_list: List[int],
):
    """Writes values to strings, depending on the type

    Parameters
    ----------
    worksheet : Worksheet
        Work sheet
    data : list
        A list with the data that is entered in the rows
    numbers_list : list
        A list with the rows numbers in which the data is entered

    Raises
    ------
    ValueError
        If one of the parameters does not match its type

    """

    if not isinstance(worksheet, Worksheet):
        raise ValueError("Parameter worksheet should be a Worksheet")
    elif not isinstance(data, list):
        raise ValueError("Parameter data should be a list")
    elif not isinstance(numbers_list, list):
        raise ValueError("Parameter numbers_list should be a list")

    for index_item, item in enumerate(data):
        row_index = numbers_list[index_item]
        if item["type"] == "filled":
            insert_filled_row(
                worksheet,
                row_index,
                str(index_item + 1),
                item["data"]["name"],
                item["data"]["units"],
                item["data"]["count"],
                item["data"]["price"],
                item["data"]["total"],
            )
        elif item["type"] == "subtitle":
            insert_merged_row(
                worksheet,
                row_index,
                str(index_item + 1),
                item["data"]["title"]
            )
        elif item["type"] == "total":
            insert_merged_row(
                worksheet,
                row_index,
                str(index_item + 1),
                item["data"]["value"]
            )


def insert_filled_row(
    worksheet: Worksheet,
    row_index: int,
    index: str,
    name: str,
    units: str,
    count: str,
    price: str,
    total: str,
) -> None:
    """Will enter data in the row

    Parameters
    ----------
    worksheet : Worksheet
        Work sheet
    row_index : ine
         The number of the row whose cells need to be merged
    index : str
        The value to be entered in the first column
    name, units, count, price, total : str
        Values to be entered in the cells

    Raises
    ------
    ValueError
        If one of the parameters does not match its type

    """

    if not isinstance(worksheet, Worksheet):
        raise ValueError("Parameter worksheet should be a Worksheet")
    elif not isinstance(row_index, int):
        raise ValueError("Parameter row_index should be a int")
    elif not isinstance(index, str):
        raise ValueError("Parameter index should be a str")
    elif not isinstance(name, str):
        raise ValueError("Parameter name should be a str")
    elif not isinstance(units, str):
        raise ValueError("Parameter units should be a str")
    elif not isinstance(count, str):
        raise ValueError("Parameter count should be a str")
    elif not isinstance(price, str):
        raise ValueError("Parameter price should be a str")
    elif not isinstance(total, str):
        raise ValueError("Parameter total should be a str")

    items = [index, name, units, count, price, total]
    cell_address = ["A", "B", "C", "D", "E", "F"]
    for i in range(6):
        insert_formatted_text(
            worksheet[f"{cell_address[i]}{row_index}"],
            items[i],
            bold=True,
            alignment="center",
        )


def insert_merged_row(
    worksheet: Worksheet,
    row_index: int,
    index: str,
    value: str,
) -> None:
    """Creates a merged cell and enters the value there

    Parameters
    ----------
    worksheet : Worksheet
        Work sheet
    row_index : ine
         The number of the row whose cells need to be merged
    index : str
        The value to be entered in the first column
    value : str
        The value to be entered in the merged cell


    Raises
    ------
    ValueError
        If one of the parameters does not match its type

    """

    if not isinstance(worksheet, Worksheet):
        raise ValueError("Parameter worksheet should be a Worksheet")
    elif not isinstance(row_index, int):
        raise ValueError("Parameter row_index should be a int")
    elif not isinstance(index, str):
        raise ValueError("Parameter index should be a str")
    elif not isinstance(value, str):
        raise ValueError("Parameter value should be a str")

    insert_formatted_text(
        worksheet[f"A{row_index}"],
        index,
        bold=True,
        alignment="center",
    )

    worksheet.merge_cells(f"B{row_index}:F{row_index}")
    insert_formatted_text(
        worksheet[f"B{row_index}"],
        value,
        bold=True,
        alignment="center",
    )


def insert_formatted_text(
    cell: Cell,
    value: str,
    bold: bool = False,
    alignment: Literal["left", "center", "right"] = "left",
    border: Border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    ),
) -> None:
    """Formats the cell text

    Parameters
    ----------
    cell : Cell
        An object of the Cell class
    value : str
        The text to be written in the cell
    bold : bool, default: False
        Text weight
    alignment : {"left", "center", "right"}, default: "left"
        Text alignment
    border : Border, default: Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )
        Cell border

    Raises
    ------
    ValueError
        If one of the parameters does not match its type

    """

    # TODO values check
    if not isinstance(cell, Cell):
        raise ValueError("Parameter cell should be a Cell")
    elif not isinstance(value, str):
        raise ValueError("Parameter value should be a str")
    elif not isinstance(bold, bool):
        raise ValueError("Parameter bold should be a bool")
    elif not isinstance(alignment, str):
        raise ValueError("Parameter alignment should be a str")
    elif not isinstance(border, Border):
        raise ValueError("Parameter border should be a Border")

    cell.value = value

    # bold
    cell.font = Font(bold=True)

    # alignment
    cell.alignment = Alignment(horizontal=alignment)

    # border
    cell.border = border
